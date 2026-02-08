"""XLSX解析器"""

import io
from typing import Dict, Any, Optional, Union
import openpyxl

from hos_m2f.model.universal_model import UniversalDocumentModel
from hos_m2f.parsers.base_parser import BaseParser


class XLSXParser(BaseParser):
    """XLSX解析器"""
    
    def parse(self, content: Union[str, bytes], options: Optional[Dict[str, Any]] = None) -> UniversalDocumentModel:
        """解析XLSX内容为中间模型
        
        Args:
            content: XLSX内容
            options: 解析选项
            
        Returns:
            UniversalDocumentModel: 通用文档模型
        """
        options = options or {}
        
        # 打开XLSX文件
        if isinstance(content, bytes):
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        else:
            workbook = openpyxl.load_workbook(content, data_only=True)
        
        # 创建通用文档模型
        model = UniversalDocumentModel()
        
        # 1. 提取元数据
        meta = self._extract_meta(workbook)
        model.set_meta(meta)
        
        # 2. 生成HTML内容
        html_content = self._generate_html_content(workbook, options)
        model.set_html_content(html_content)
        
        # 3. 分析文档结构
        structure = self._analyze_structure(workbook)
        for item in structure:
            model.add_structure_item(item)
        
        # 4. 提取语义信息
        semantics = self._extract_semantics(workbook)
        model.get_json()["semantics"].update(semantics)
        
        return model
    
    def _extract_meta(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """提取元数据
        
        Args:
            workbook: Excel工作簿
            
        Returns:
            Dict[str, Any]: 元数据字典
        """
        meta = {
            "title": "Excel Document",
            "author": "",
            "description": f"Excel document with {len(workbook.sheetnames)} sheets",
            "tags": ["Excel", "Spreadsheet"],
            "sheet_count": len(workbook.sheetnames)
        }
        
        # 尝试从工作簿属性中提取元数据
        try:
            properties = workbook.properties
            if properties.title:
                meta["title"] = properties.title
            if properties.creator:
                meta["author"] = properties.creator
            if properties.description:
                meta["description"] = properties.description
        except Exception:
            pass
        
        return meta
    
    def _generate_html_content(self, workbook: openpyxl.Workbook, options: Dict[str, Any]) -> str:
        """生成HTML内容
        
        Args:
            workbook: Excel工作簿
            options: 生成选项
            
        Returns:
            str: HTML内容
        """
        html_parts = []
        
        # 为每个工作表生成HTML表格
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            html_parts.append(f"<h2>{sheet_name}</h2>")
            html_parts.append(self._sheet_to_html(worksheet, options))
        
        return "<div class='excel-content'>" + "\n".join(html_parts) + "</div>"
    
    def _sheet_to_html(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, options: Dict[str, Any]) -> str:
        """将工作表转换为HTML表格
        
        Args:
            worksheet: Excel工作表
            options: 转换选项
            
        Returns:
            str: HTML表格
        """
        html = "<table border='1' cellpadding='4' cellspacing='0' style='border-collapse: collapse;'>"
        
        # 获取数据范围
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # 生成表头
        html += "<thead><tr>"
        for col in range(1, max_col + 1):
            header_cell = worksheet.cell(row=1, column=col)
            header_value = header_cell.value or ""
            html += f"<th>{header_value}</th>"
        html += "</tr></thead>"
        
        # 生成表格内容
        html += "<tbody>"
        for row in range(2, max_row + 1):
            html += "<tr>"
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell_value = cell.value or ""
                html += f"<td>{cell_value}</td>"
            html += "</tr>"
        html += "</tbody>"
        
        html += "</table>"
        return html
    
    def _analyze_structure(self, workbook: openpyxl.Workbook) -> list:
        """分析文档结构
        
        Args:
            workbook: Excel工作簿
            
        Returns:
            list: 结构项列表
        """
        structure = []
        
        for i, sheet_name in enumerate(workbook.sheetnames):
            worksheet = workbook[sheet_name]
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            structure.append({
                "type": "sheet",
                "name": sheet_name,
                "position": i,
                "rows": max_row,
                "columns": max_col
            })
            
            structure.append({
                "type": "table",
                "sheet": sheet_name,
                "rows": max_row,
                "columns": max_col
            })
        
        return structure
    
    def _extract_semantics(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """提取语义信息
        
        Args:
            workbook: Excel工作簿
            
        Returns:
            Dict[str, Any]: 语义信息字典
        """
        semantics = {
            "domain_tag": "spreadsheet",
            "sheet_info": {},
            "data_range": {}
        }
        
        # 提取工作表信息
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            semantics["sheet_info"][sheet_name] = {
                "rows": max_row,
                "columns": max_col,
                "cell_count": max_row * max_col
            }
        
        return semantics
